## Populates an EXCEL file with the information from the models in the relevant folders


import os
import random
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ======================================================
# 🔧 CONFIGURATION
# ======================================================

EXCEL_PATH = r"E:\gen\civitai\i2v\list.xlsx"

MODEL_PATHS = {
    "lora":       r"E:/gen/img/Models/loras",
    "checkpoint": r"E:/gen/img/Models/checkpoints",
    "workflow":   r"E:/gen/img/user/default/workflows",
    "working-images":      r"E:\gen\civitai\i2v\added",
    "not-working-images":  r"E:\gen\civitai\i2v\doesntwork",
    "results":             r"E:\gen\img\Outputs"
}

BASE_MODEL_PATTERNS = {
    "ILLUSTRIOUS":  r"illustrious[a-z0-9_-]*",
    "WAN":          r"wan[a-z0-9_-]*",
    "HUNYUAN":      r"hunyuan[a-z0-9_-]*",
    "QWEN":         r"qwen[a-z0-9_-]*",
    "HiDream":      r"hidream[a-z0-9_-]*",
    "SDXL":         r"sd[\s\-_.]*xl",
    "SD1.5":        r"sd[\s\-_.]*1[\s\-_.]*5",
    "SD3":          r"sd[\s\-_.]*3",
    "PONY":         r"pony[a-z0-9_-]*",
    "FLUX":         r"flux[a-z0-9_-]*"
}

SKIP_EXISTING_FILES = True

HEADERS = [
    "ID", "Filename", "Model Type", "Base Model", "Size",
    "Trigger Words", "Download Link", "Description"
]

# ======================================================
# 🧱 HELPERS
# ======================================================

def generate_id(existing_ids):
    while True:
        new_id = f"A{random.randint(1000, 9999)}"
        if new_id not in existing_ids:
            return new_id

def get_file_size_gb(filepath):
    size_bytes = os.path.getsize(filepath)
    size_gb = size_bytes / (1024 ** 3)
    return f"{size_gb:.1f}GB"

def detect_base_model(filename):
    lower_name = filename.lower()
    for label, pattern in BASE_MODEL_PATTERNS.items():
        if re.search(pattern, lower_name, re.IGNORECASE):
            return label
    return ""

def load_or_create_excel(path):
    """Open existing workbook or create a new one (no headers yet)."""
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
    return wb, ws

def append_header_block(ws, folder_name):
    """Append the 3-row header block for a folder."""
    start_row = ws.max_row + 1
    last_col = len(HEADERS)
    col_letter = get_column_letter(last_col)

    # Row 1: black separator
    ws.append([""] * last_col)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=last_col)
    sep = ws.cell(row=start_row, column=1)
    sep.fill = PatternFill(fill_type="solid", fgColor="FF000000")  # black
    ws.row_dimensions[start_row].height = 8

    # Row 2: dark-green title
    ws.append([""] * last_col)
    ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=last_col)
    title = ws.cell(row=start_row + 1, column=1)
    title.value = os.path.basename(folder_name.rstrip("\\/")) or folder_name
    title.font = Font(bold=True, size=18, color="FFFFFFFF")
    title.fill = PatternFill(fill_type="solid", fgColor="FF006400")  # dark green
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Row 3: light-green column headers
    ws.append(HEADERS)
    for i in range(1, last_col + 1):
        c = ws.cell(row=start_row + 2, column=i)
        c.font = Font(bold=True, size=15)
        c.fill = PatternFill(fill_type="solid", fgColor="FF90EE90")
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust widths once (optional)
    widths = [14, 48, 16, 16, 10, 28, 36, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def append_folder_data(ws, model_type, folder, existing_files, existing_ids):
    """Write file rows under the current header block."""
    files = [f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
    print(f"📂 Found {len(files)} files in {folder}")

    for filename in files:
        if SKIP_EXISTING_FILES and filename in existing_files:
            print(f"⏭️ Skipping existing file: {filename}")
            continue

        filepath = os.path.join(folder, filename)
        entry_id = generate_id(existing_ids)
        existing_ids.add(entry_id)
        size = get_file_size_gb(filepath)
        base_model = detect_base_model(filename)

        ws.append([
            entry_id,
            filename,
            model_type,
            base_model,
            size,
            "", "", ""
        ])

        existing_files.add(filename)
        print(f"➕ Added {entry_id}: {filename} ({model_type}, {size}, base={base_model or 'none'})")

# ======================================================
# 🚀 MAIN
# ======================================================

def main():
    print("=== Starting Excel Model Index ===")

    wb, ws = load_or_create_excel(EXCEL_PATH)
    existing_files = set()
    existing_ids = set()

    for model_type, folder in MODEL_PATHS.items():
        print(f"\n🔍 Scanning {model_type}: {folder}")

        if not os.path.exists(folder):
            print(f"❌ Directory not found: {folder}")
            continue

        append_header_block(ws, folder)
        append_folder_data(ws, model_type, folder, existing_files, existing_ids)

    wb.save(EXCEL_PATH)
    print(f"\n✅ Excel registry updated at {EXCEL_PATH}")
    print("=== Done ===")

# ======================================================
# 🏁 RUN
# ======================================================
if __name__ == "__main__":
    main()
